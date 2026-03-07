"""
template_filler.py
==================
Generates the standard voyage performance report in Excel format.

Layout matches the standard template (standard_template.md):
  - One sheet per voyage leg
  - Section 1: Voyage Information
  - Section 2: Speed Warranty (Clause 24.4)
  - Section 4: GCU Compliance (Clause 23.5(b))
  - Section 7: Segment Data for Fuel Consumption (Clause 24.5)

Column layout per voyage sheet:
  A = Parameter labels (width 31)
  B = Segment 1 data
  C = Segment 2 data (or Total if single segment)
  ...additional segments...
  Then: Total, Units, Notes, TCP Ref
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# =============================================================================
# Style constants
# =============================================================================

_TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)

_HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
_HEADER_FONT = Font(bold=True, size=10)

_COL_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_COL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

_DATA_FILL = PatternFill("solid", fgColor="D6EAF8")
_DATA_FONT = Font(color="1F4E79", size=10)

_SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
_SUBHEADER_FONT = Font(bold=True, size=10)

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_NUM_2DP = "0.00"
_NUM_3DP = "0.000"
_NUM_4DP = "0.0000"
_NUM_1DP = "0.0"


# =============================================================================
# Helper: write a cell with optional styling
# =============================================================================

def _w(ws, row, col, value, font=None, fill=None, align=None, fmt=None):
    """Write a value to a cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell


def _write_row(ws, row, col_start, values, font=None, fill=None, align=None, fmt=None):
    """Write a list of values to consecutive cells in a row."""
    for i, v in enumerate(values):
        _w(ws, row, col_start + i, v, font=font, fill=fill, align=align, fmt=fmt)


# =============================================================================
# Section 1: VOYAGE INFORMATION
# =============================================================================

def _write_section1(ws, metadata: dict, n_seg: int):
    """
    Write Section 1: Voyage Information (rows 3-15).
    metadata should contain: voyage_no, voyage_type, fuel_mode,
    load_port, discharge_port, distance, dep_datetime, arr_datetime,
    cargo_density, lcv, charter_year
    """
    # Section title
    _w(ws, 3, 1, "1. VOYAGE INFORMATION", _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)
    # Merge title across columns
    last_col = 1 + n_seg + 4  # A + segments + Total/Units/Notes/TCPRef
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

    # Header row
    _w(ws, 4, 1, "Field", _HEADER_FONT, _HEADER_FILL, _ALIGN_LEFT)
    _w(ws, 4, 2, "Value", _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 4, 3, "Units", _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 4, 4, "Notes", _HEADER_FONT, _HEADER_FILL, _ALIGN_LEFT)

    fields = [
        (5, "Voyage Number", metadata.get("voyage_no", ""), "", ""),
        (6, "Voyage Type", metadata.get("voyage_type", ""), "", "Laden/Ballast"),
        (7, "Fuel Mode", metadata.get("fuel_mode", ""), "", ""),
        (8, "Load Port", metadata.get("load_port", ""), "", ""),
        (9, "Discharge Port", metadata.get("discharge_port", ""), "", ""),
        (10, "Total Distance",
         metadata.get("distance", ""), "nm",
         f'{metadata.get("duration_days", 0):.1f} days' if metadata.get("duration_days") else ""),
        (11, "CTMS Open (Departure)",
         str(metadata.get("dep_datetime", ""))[:16], "", "UTC"),
        (12, "CTMS Close (Arrival)",
         str(metadata.get("arr_datetime", ""))[:16], "", "UTC"),
        (13, "LNG Density",
         metadata.get("cargo_density", ""), "mts/cbm", ""),
        (14, "LNG LCV",
         metadata.get("lcv", ""), "MJ/kg", ""),
        (15, "Charter Year",
         metadata.get("charter_year", 1), "", ""),
    ]

    for row, label, value, unit, note in fields:
        _w(ws, row, 1, label, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, value, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER,
           fmt=_NUM_2DP if isinstance(value, float) else None)
        _w(ws, row, 3, unit, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
        _w(ws, row, 4, note, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)


# =============================================================================
# Section 2: SPEED WARRANTY
# =============================================================================

def _write_section2(ws, totals: dict, n_seg: int):
    """
    Write Section 2: Speed Warranty (rows 17-28).
    """
    last_col = 1 + n_seg + 4
    _w(ws, 17, 1, "2. SPEED WARRANTY (Clause 24.4)", _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)
    ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=last_col)

    # Header
    _w(ws, 18, 1, "Exclusion Type", _HEADER_FONT, _HEADER_FILL, _ALIGN_LEFT)
    _w(ws, 18, 2, "Hours", _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 18, 3, "Notes", _HEADER_FONT, _HEADER_FILL, _ALIGN_LEFT)

    _w(ws, 19, 1, "Exclusions:", _SUBHEADER_FONT, _SUBHEADER_FILL, _ALIGN_LEFT)

    exclusion_rows = [
        (20, "Weather >BF5 >6hrs",
         totals.get("weather_excl_hours", 0)),
        (21, "Poor visibility", 0),
        (22, "Congested waters", 0),
        (23, "Typhoon avoidance", 0),
        (24, "Charterer stops", 0),
        (25, "Off-hire", 0),
        (26, "Regulatory", 0),
        (27, "Save life/property", 0),
    ]

    for row, label, value in exclusion_rows:
        _w(ws, row, 1, label, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, value, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, fmt=_NUM_2DP)

    # Total exclusions
    total_excl = totals.get("total_excl_hours", 0)
    _w(ws, 28, 1, "Total Exclusions", Font(bold=True, color="1F4E79"), _DATA_FILL, _ALIGN_LEFT)
    _w(ws, 28, 2, total_excl, Font(bold=True, color="1F4E79"), _DATA_FILL, _ALIGN_CENTER,
       fmt=_NUM_2DP)


# =============================================================================
# Section 4: GCU COMPLIANCE
# =============================================================================

def _write_section4(ws, totals: dict, n_seg: int):
    """
    Write Section 4: GCU Compliance (rows 30-39).
    """
    last_col = 1 + n_seg + 4
    _w(ws, 30, 1, "4. GCU COMPLIANCE (Clause 23.5(b))", _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)
    ws.merge_cells(start_row=30, start_column=1, end_row=30, end_column=last_col)

    # Header
    _w(ws, 31, 1, "Parameter", _HEADER_FONT, _HEADER_FILL, _ALIGN_LEFT)
    _w(ws, 31, 2, "Value", _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 31, 3, "Units", _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)

    gcu_used = totals.get("gcu_used", False)
    gcu_total = totals.get("gcu_total", 0)

    # Above 12 knots
    _w(ws, 32, 1, "GCU ABOVE 12 KNOTS", _SUBHEADER_FONT, _SUBHEADER_FILL, _ALIGN_LEFT)
    rows_above = [
        (33, "GCU Used Above 12kn?", "Yes" if gcu_used else "No", ""),
        (34, "GCU Volume Above 12kn", gcu_total if gcu_used else 0, "cbm"),
        (35, "Excess GCU (Above)", 0, "cbm"),
    ]
    for row, label, value, unit in rows_above:
        _w(ws, row, 1, label, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, value, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER,
           fmt=_NUM_2DP if isinstance(value, (int, float)) else None)
        _w(ws, row, 3, unit, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)

    # Below 12 knots
    _w(ws, 36, 1, "GCU BELOW 12 KNOTS", _SUBHEADER_FONT, _SUBHEADER_FILL, _ALIGN_LEFT)
    rows_below = [
        (37, "Below 12kn Duration", 0, "days"),
        (38, "Qty burned in GCU", 0, "cbm"),
        (39, "Authorized/Excess GCU", 0, "cbm"),
    ]
    for row, label, value, unit in rows_below:
        _w(ws, row, 1, label, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, value, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, fmt=_NUM_2DP)
        _w(ws, row, 3, unit, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)


# =============================================================================
# Section 7: SEGMENT DATA FOR FUEL CONSUMPTION
# =============================================================================

def _write_section7(ws, segments: list[dict], totals: dict, n_seg: int):
    """
    Write Section 7: Segment Data for Fuel Consumption (rows 41+).

    Columns: A=labels, B..=Seg1,Seg2..., then Total, Units, Notes, TCP Ref
    """
    last_col = 1 + n_seg + 4  # A + segments + Total/Units/Notes/TCPRef
    total_col = 1 + n_seg + 1
    units_col = total_col + 1
    notes_col = units_col + 1
    tcp_col = notes_col + 1

    # --- Section Title (row 41) ---
    _w(ws, 41, 1, "7. SEGMENT DATA FOR FUEL CONSUMPTION (Clause 24.5)",
       _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)
    ws.merge_cells(start_row=41, start_column=1, end_row=41, end_column=last_col)

    # --- Column Headers (row 42) ---
    _w(ws, 42, 1, "Parameter", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_LEFT)
    for s in range(n_seg):
        _w(ws, 42, 2 + s, f"Segment {s + 1}", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 42, total_col, "Total", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 42, units_col, "Units", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 42, notes_col, "Notes", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)
    _w(ws, 42, tcp_col, "TCP Ref", _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)

    # Helper to write a data row across segments + total
    def seg_row(row, label, key, unit="", note="", tcp="", fmt=_NUM_2DP,
                is_subheader=False, total_val=None):
        if is_subheader:
            _w(ws, row, 1, label, _SUBHEADER_FONT, _SUBHEADER_FILL, _ALIGN_LEFT)
            for c in range(2, last_col + 1):
                _w(ws, row, c, "", _SUBHEADER_FONT, _SUBHEADER_FILL)
            return

        _w(ws, row, 1, label, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)

        for s in range(n_seg):
            val = segments[s].get(key) if key else None
            _w(ws, row, 2 + s, val, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER,
               fmt=fmt if isinstance(val, (int, float)) else None)

        # Total column
        tv = total_val if total_val is not None else (totals.get(key) if key else None)
        _w(ws, row, total_col, tv, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER,
           fmt=fmt if isinstance(tv, (int, float)) else None)

        _w(ws, row, units_col, unit, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
        _w(ws, row, notes_col, note, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, tcp_col, tcp, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)

    # --- Basic Segment Info (rows 43-49) ---
    seg_row(43, "BASIC SEGMENT INFO", None, is_subheader=True)

    # Start/End datetime — use custom writing for string values
    seg_row(44, "Start Date/Time", "start_datetime", "", "", "", fmt=None)
    seg_row(45, "End Date/Time", "end_datetime", "", "", "", fmt=None)

    # Fix datetime display: overwrite with string format
    for s in range(n_seg):
        dt = segments[s].get("start_datetime")
        _w(ws, 44, 2 + s, str(dt)[:16] if dt else "",
           _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
        dt = segments[s].get("end_datetime")
        _w(ws, 45, 2 + s, str(dt)[:16] if dt else "",
           _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
    # Total column for dates
    _w(ws, 44, total_col, str(totals.get("start_datetime", ""))[:16],
       _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
    _w(ws, 45, total_col, str(totals.get("end_datetime", ""))[:16],
       _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)

    seg_row(46, "Duration", "duration_days", "days", "", "", fmt=_NUM_4DP)
    seg_row(47, "Distance", "distance", "nm", "", "")
    seg_row(48, "Instructed Speed", "instructed_speed", "knots", "", "")
    seg_row(49, "Fuel Mode", "fuel_mode", "", "", "", fmt=None)

    # --- Exclusions Per Segment (rows 51-55) ---
    seg_row(51, "EXCLUSIONS PER SEGMENT", None, is_subheader=True)
    seg_row(52, "Weather BF>5 >6hrs", "weather_bf5_hours", "hours")
    seg_row(53, "Weather Excl Hours", "weather_excl_hours", "hours")
    seg_row(54, "Other Exclusion Hours", "other_excl_hours", "hours")
    seg_row(55, "Total Excl Hours", "total_excl_hours", "hours")

    # --- Speed Exclusions (rows 57-63) ---
    seg_row(57, "SPEED EXCLUSIONS", None, is_subheader=True)
    seg_row(58, "Weather Hours", "weather_excl_hours", "hours")
    seg_row(59, "Weather Distance", "weather_excl_distance", "nm")
    seg_row(60, "Regulatory Hours", "regulatory_excl_hours", "hours")
    seg_row(61, "Regulatory Distance", "regulatory_excl_distance", "nm")
    seg_row(62, "Total Excl Hours", "total_speed_excl_hours", "hours")
    seg_row(63, "Total Excl Distance", "total_speed_excl_distance", "nm")

    # --- Calculated Speed (rows 65-69) ---
    seg_row(65, "CALCULATED SPEED (Clause 24.5(d))", None, is_subheader=True)
    seg_row(66, "Net Duration", "net_duration_days", "days", "", "", fmt=_NUM_4DP)
    seg_row(67, "Net Distance", "net_distance", "nm")
    seg_row(68, "Actual Avg Speed", "actual_avg_speed", "knots",
            "NetDist/(NetDur×24)", "24.5(d)", fmt=_NUM_3DP)
    seg_row(69, "Reference Speed", "reference_speed", "knots",
            "MIN(Actual, Instructed)", "", fmt=_NUM_3DP)

    # --- Actual Fuel Consumed (rows 71-81) ---
    seg_row(71, "ACTUAL FUEL CONSUMED", None, is_subheader=True)
    seg_row(72, "LNG Consumed", "lng_consumed", "cbm", "", "24.5(h)")
    seg_row(73, "MGO Consumed (prop+pilot+blr)", "mgo_consumed", "mts", "", "")
    seg_row(74, "VLSFO Consumed (prop+pilot+blr)", "vlsfo_consumed", "mts", "", "")
    # row 75 blank
    seg_row(76, "MGO Pilot", "mgo_pilot", "mts")
    seg_row(77, "MGO Boiler", "mgo_boiler", "mts")
    seg_row(78, "MGO Propulsion", "mgo_propulsion", "mts", "Total-Pilot-Boiler")
    seg_row(79, "VLSFO Pilot", "vlsfo_pilot", "mts")
    seg_row(80, "VLSFO Boiler", "vlsfo_boiler", "mts")
    seg_row(81, "VLSFO Propulsion", "vlsfo_propulsion", "mts", "Total-Pilot-Boiler")

    # --- Fuel Consumption Exclusions (rows 83-95) ---
    seg_row(83, "FUEL CONSUMPTION EXCLUSIONS", None, is_subheader=True)
    seg_row(84, "Gas Fuel in Excluded Periods", None, is_subheader=True)
    seg_row(85, "LNG in Weather Excl", "excl_lng_weather", "cbm")
    seg_row(86, "LNG in Other Excl", "excl_lng_other", "cbm")
    seg_row(87, "Total LNG Excluded", "excl_lng_total", "cbm")
    # row 88 blank
    seg_row(89, "Liquid Fuel in Excluded Periods", None, is_subheader=True)
    seg_row(90, "MGO in Weather Excl", "excl_mgo_weather", "mts")
    seg_row(91, "MGO in Other Excl", "excl_mgo_other", "mts")
    seg_row(92, "Total MGO Excluded", "excl_mgo_total", "mts")
    seg_row(93, "VLSFO in Weather Excl", "excl_vlsfo_weather", "mts")
    seg_row(94, "VLSFO in Other Excl", "excl_vlsfo_other", "mts")
    seg_row(95, "Total VLSFO Excluded", "excl_vlsfo_total", "mts")

    # --- Net Fuel Quantities (rows 97-100) ---
    seg_row(97, "NET FUEL QUANTITIES", None, is_subheader=True)
    seg_row(98, "Net LNG", "net_lng", "cbm", "Actual - Excluded")
    seg_row(99, "Net MGO", "net_mgo", "mts", "Actual - Excluded")
    seg_row(100, "Net VLSFO", "net_vlsfo", "mts", "Actual - Excluded")

    # --- LCV Values (rows 102-104) ---
    seg_row(102, "LCV VALUES", None, is_subheader=True)

    # LCV and density: write from metadata (same across all segments typically)
    # Use the first segment's values or from auxiliary data
    _w(ws, 103, 1, "LNG LCV", _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
    _w(ws, 104, 1, "LNG Density", _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
    _w(ws, 103, units_col, "MJ/kg", _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
    _w(ws, 104, units_col, "mts/cbm", _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)

    # --- Reliquefaction Data (rows 106-108) ---
    seg_row(106, "RELIQUEFACTION DATA", None, is_subheader=True)
    seg_row(107, "Subcooler/Reliq Hours", "reliq_hours", "hours")
    seg_row(108, "Subcooler/Reliq Avg Load", "reliq_avg_load", "%", "", "", fmt=_NUM_1DP)

    # --- ECA Data (rows 110-112) ---
    seg_row(110, "ECA DATA", None, is_subheader=True)
    seg_row(111, "ECA Hours", None, "hours")
    seg_row(112, "ECA Distance", None, "nm")

    # --- Remarks (rows 114+) ---
    seg_row(114, "REMARKS", None, is_subheader=True)

    # Write remarks per segment
    max_remarks = 0
    for s in range(n_seg):
        remarks = segments[s].get("remarks", [])
        max_remarks = max(max_remarks, len(remarks))

    for r_idx in range(max_remarks):
        row = 115 + r_idx
        _w(ws, row, 1, f"Day {r_idx + 1}", _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        for s in range(n_seg):
            remarks = segments[s].get("remarks", [])
            val = remarks[r_idx] if r_idx < len(remarks) else ""
            _w(ws, row, 2 + s, val, _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)


# =============================================================================
# Section: Speed Anomaly Detection  (Rule A)
# =============================================================================

def _write_speed_anomalies(ws, anomalies: list[dict], start_row: int) -> int:
    """
    Write speed anomaly alerts after the main data sections.
    Returns the next available row.
    """
    _w(ws, start_row, 1, "SPEED ANOMALY DETECTION (Rule A)",
       _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)

    if not anomalies:
        _w(ws, start_row + 1, 1, "No speed anomalies detected",
           _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        return start_row + 3

    headers = ["Date", "Reported Speed (kts)", "Weighted Avg (kts)", "Report Type"]
    for i, h in enumerate(headers):
        _w(ws, start_row + 1, 1 + i, h, _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)

    for j, a in enumerate(anomalies):
        row = start_row + 2 + j
        _w(ws, row, 1, str(a.get("datetime", ""))[:16], _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, a.get("avg_speed", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)
        _w(ws, row, 3, a.get("weighted_avg", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_3DP)
        _w(ws, row, 4, a.get("report_type", ""), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)

    return start_row + 2 + len(anomalies) + 1


# =============================================================================
# Section: Intermediate Stops / Bunkering  (Rule B)
# =============================================================================

def _write_intermediate_stops(ws, stops: list[dict], start_row: int) -> int:
    """
    Write mid-voyage bunkering/port call details.
    Returns the next available row.
    """
    _w(ws, start_row, 1, "MID-VOYAGE PORT CALLS / BUNKERING (Rule B)",
       _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)

    if not stops:
        _w(ws, start_row + 1, 1, "No intermediate stops detected",
           _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        return start_row + 3

    headers = ["Port", "Arrival", "Departure", "Duration (hrs)",
               "LNG (m\u00b3)", "MGO (MT)", "VLSFO (MT)"]
    for i, h in enumerate(headers):
        _w(ws, start_row + 1, 1 + i, h, _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER)

    for j, stop in enumerate(stops):
        row = start_row + 2 + j
        _w(ws, row, 1, stop.get("port_name", ""), _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        _w(ws, row, 2, stop.get("arr_datetime", ""), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
        _w(ws, row, 3, stop.get("dep_datetime", ""), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER)
        _w(ws, row, 4, stop.get("duration_hours", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)
        _w(ws, row, 5, stop.get("lng_consumed", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)
        _w(ws, row, 6, stop.get("mgo_consumed", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)
        _w(ws, row, 7, stop.get("vlsfo_consumed", 0), _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)

    return start_row + 2 + len(stops) + 1


# =============================================================================
# AI Review Sheet
# =============================================================================

_SEVERITY_FILLS = {
    "error":   PatternFill("solid", fgColor="F4CCCC"),
    "warning": PatternFill("solid", fgColor="FFF2CC"),
    "info":    PatternFill("solid", fgColor="D9EAD3"),
}

_SEVERITY_FONTS = {
    "error":   Font(bold=True, color="CC0000", size=10),
    "warning": Font(bold=True, color="7F6000", size=10),
    "info":    Font(color="274E13", size=10),
}


def create_ai_review_sheet(wb: Workbook, all_alerts: list[dict]):
    """
    Create an 'AI Review' sheet with all alerts from the AI analyst.

    Parameters
    ----------
    wb         : Workbook to add the sheet to
    all_alerts : list of {voyage_no, severity, category, message, details}
    """
    ws = wb.create_sheet("AI Review")

    # Title
    _w(ws, 1, 1, "AI ANALYST REVIEW", _TITLE_FONT, _TITLE_FILL, _ALIGN_LEFT)

    if not all_alerts:
        _w(ws, 3, 1, "No issues found — all checks passed.",
           _DATA_FONT, _DATA_FILL, _ALIGN_LEFT)
        ws.column_dimensions["A"].width = 50
        return

    # Headers
    headers = ["Voyage", "Severity", "Category", "Finding", "Details"]
    for i, h in enumerate(headers):
        _w(ws, 3, 1 + i, h, _COL_HEADER_FONT, _COL_HEADER_FILL, _ALIGN_CENTER)

    # Sort: errors first, then warnings, then info
    severity_order = {"error": 0, "warning": 1, "info": 2}
    sorted_alerts = sorted(
        all_alerts, key=lambda a: severity_order.get(a.get("severity", "info"), 9)
    )

    for j, alert in enumerate(sorted_alerts):
        row = 4 + j
        sev = alert.get("severity", "info")
        fill = _SEVERITY_FILLS.get(sev, _DATA_FILL)
        font = _SEVERITY_FONTS.get(sev, _DATA_FONT)

        _w(ws, row, 1, alert.get("voyage_no", ""), font, fill, _ALIGN_CENTER)
        _w(ws, row, 2, sev.upper(), font, fill, _ALIGN_CENTER)
        _w(ws, row, 3, alert.get("category", ""), font, fill, _ALIGN_CENTER)
        _w(ws, row, 4, alert.get("message", ""), font, fill, _ALIGN_LEFT)
        _w(ws, row, 5, alert.get("details", ""), font, fill, _ALIGN_LEFT)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60

    ws.freeze_panes = "A4"

    logger.info("AI Review sheet created with %d alert(s)", len(all_alerts))


# =============================================================================
# Column width & formatting
# =============================================================================

def _apply_formatting(ws, n_seg: int):
    """
    Apply standard column widths and formatting.

    Widths: A=31, B-C=16, then remaining data cols=16,
    Total=23, Units=14, Notes=16.5, TCP Ref=15
    """
    total_col = 1 + n_seg + 1
    units_col = total_col + 1
    notes_col = units_col + 1
    tcp_col = notes_col + 1

    # Column A = labels
    ws.column_dimensions["A"].width = 31

    # Segment columns
    for s in range(n_seg):
        col_letter = get_column_letter(2 + s)
        ws.column_dimensions[col_letter].width = 16

    # Total, Units, Notes, TCP Ref
    ws.column_dimensions[get_column_letter(total_col)].width = 23
    ws.column_dimensions[get_column_letter(units_col)].width = 14
    ws.column_dimensions[get_column_letter(notes_col)].width = 16.5
    ws.column_dimensions[get_column_letter(tcp_col)].width = 15

    # Freeze panes at row 42 (column headers)
    ws.freeze_panes = "B43"


# =============================================================================
# Public API
# =============================================================================

def create_voyage_sheet(
    wb: Workbook,
    sheet_name: str,
    segments: list[dict],
    totals: dict,
    metadata: dict,
    auxiliary: dict | None = None,
    speed_anomalies: list[dict] | None = None,
    intermediate_stops: list[dict] | None = None,
):
    """
    Create a single voyage sheet with all standard sections.

    Parameters
    ----------
    wb                  : Workbook to add the sheet to
    sheet_name          : Tab name (e.g., "Voyage 11")
    segments            : list[dict] from calculator.compute_all_segments()["segments"]
    totals              : dict from calculator.compute_all_segments()["totals"]
    metadata            : dict with voyage info (voyage_no, voyage_type, ports, etc.)
    auxiliary           : dict from data_extractor.extract_auxiliary()
    speed_anomalies     : list[dict] from calculator — Rule A flagged rows
    intermediate_stops  : list[dict] from data_extractor — Rule B stop data
    """
    ws = wb.create_sheet(sheet_name)
    n_seg = len(segments)

    if n_seg == 0:
        _w(ws, 1, 1, "No segment data available")
        return

    # Write all standard sections
    _write_section1(ws, metadata, n_seg)
    _write_section2(ws, totals, n_seg)
    _write_section4(ws, totals, n_seg)
    _write_section7(ws, segments, totals, n_seg)

    # Write LCV/density from auxiliary data
    if auxiliary:
        total_col = 1 + n_seg + 1
        lcv_val = auxiliary.get("lcv")
        density_val = auxiliary.get("cargo_density")
        if lcv_val is not None:
            _w(ws, 103, 2, lcv_val, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_1DP)
            _w(ws, 103, total_col, lcv_val, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_1DP)
        if density_val is not None:
            _w(ws, 104, 2, density_val, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)
            _w(ws, 104, total_col, density_val, _DATA_FONT, _DATA_FILL, _ALIGN_CENTER, _NUM_2DP)

    # Determine next row after remarks
    max_remarks = max((len(s.get("remarks", [])) for s in segments), default=0)
    next_row = 115 + max_remarks + 2

    # Write Rule A: Speed Anomaly section
    if speed_anomalies is not None:
        next_row = _write_speed_anomalies(ws, speed_anomalies, next_row)

    # Write Rule B: Intermediate Stops section
    if intermediate_stops is not None:
        next_row = _write_intermediate_stops(ws, intermediate_stops, next_row)

    # Apply formatting
    _apply_formatting(ws, n_seg)

    logger.info(
        "Voyage sheet '%s' created with %d segments", sheet_name, n_seg
    )


def fill_template(
    output_path: str | Path,
    voyages: list[dict],
    vessel_config: dict | None = None,
):
    """
    Generate the full voyage performance report.

    Parameters
    ----------
    output_path   : Path for the output Excel file
    voyages       : list of voyage result dicts, each containing:
                    - "computed": dict from calculator.compute_all_segments()
                    - "metadata": dict with voyage info
                    - "auxiliary": dict from data_extractor.extract_auxiliary()
    vessel_config : vessel configuration dict
    """
    output_path = Path(output_path)
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    all_ai_alerts = []

    for voy in voyages:
        computed = voy["computed"]
        metadata = voy["metadata"]
        auxiliary = voy.get("auxiliary", {})

        # Sheet name from voyage number
        voyage_no = metadata.get("voyage_no", "Unknown")
        voyage_type = metadata.get("voyage_type", "")
        sheet_name = f"Voyage {voyage_no}"

        create_voyage_sheet(
            wb=wb,
            sheet_name=sheet_name,
            segments=computed["segments"],
            totals=computed["totals"],
            metadata=metadata,
            auxiliary=auxiliary,
            speed_anomalies=computed.get("speed_anomalies", []),
            intermediate_stops=metadata.get("intermediate_stops", []),
        )

        # Collect AI alerts for the review sheet
        for alert in voy.get("ai_alerts", []):
            alert["voyage_no"] = voyage_no
            all_ai_alerts.append(alert)

    # Create AI Review sheet if there are alerts
    if all_ai_alerts:
        create_ai_review_sheet(wb, all_ai_alerts)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Report saved to: %s", output_path)
    return str(output_path)
