# =============================================================================
# highlight_report.py  –  Generate a highlighted copy of the raw data Excel
# =============================================================================
# Produces a second report where:
#   1. Each voyage's rows (DEPARTURE → ARRIVAL) have a distinct background color
#   2. All columns used for report generation have RED BOLD font
# =============================================================================

from __future__ import annotations

import logging
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import COL

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Voyage background colors (light pastel shades for readability)
# ---------------------------------------------------------------------------
VOYAGE_COLORS = [
    "C6EFCE",  # Light green
    "BDD7EE",  # Light blue
    "FCE4D6",  # Light orange
    "D9D2E9",  # Light purple
    "FFF2CC",  # Light yellow
    "D5F5E3",  # Light mint
    "FADBD8",  # Light pink
    "D6EAF8",  # Light sky blue
    "E8DAEF",  # Light lavender
    "FDEBD0",  # Light peach
]

# Red bold font for columns used in report generation
RED_BOLD_FONT = Font(color="FF0000", bold=True)


def generate_highlighted_report(
    raw_excel_path: str | Path,
    output_path: str | Path,
    voyages: list[dict],
    sheet_name: str = "Sheet",
) -> None:
    """
    Create a highlighted copy of the raw Excel file.

    Parameters
    ----------
    raw_excel_path : path to the original raw noon-report Excel
    output_path    : path for the highlighted output file
    voyages        : list of voyage dicts from detect_voyages()
                     (must have 'dep_row' and 'arr_row' keys)
    sheet_name     : name of the sheet to highlight
    """
    raw_excel_path = Path(raw_excel_path)
    output_path = Path(output_path)

    # Load workbook preserving original formatting
    wb = openpyxl.load_workbook(str(raw_excel_path))

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # -- Determine which Excel columns (1-based) are used by the rules --
    # COL values are 0-based DataFrame indices; Excel columns are 1-based
    # Add +2 because: +1 for 0→1 indexing, +1 for the header row offset
    # Actually, the raw Excel is loaded with header=0 in pandas, so
    # DataFrame column index N corresponds to Excel column N+1 (1-based).
    used_col_indices_0based = set(COL.values())
    used_excel_cols = {idx + 1 for idx in used_col_indices_0based}  # 1-based

    logger.info(
        "Highlighting %d voyages and %d used columns.",
        len(voyages), len(used_excel_cols),
    )

    total_rows = ws.max_row
    total_cols = ws.max_column

    # -- 1. Apply voyage row background colors --
    for v_idx, voyage in enumerate(voyages):
        color_hex = VOYAGE_COLORS[v_idx % len(VOYAGE_COLORS)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        # dep_row and arr_row are 0-based DataFrame row indices.
        # In the Excel file, row 1 = header, row 2 = first data row.
        # So DataFrame row 0 → Excel row 2.
        excel_start = voyage["dep_row"] + 2  # +1 for 0→1, +1 for header
        excel_end = voyage["arr_row"] + 2

        for row_num in range(excel_start, excel_end + 1):
            for col_num in range(1, total_cols + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = fill

        logger.debug(
            "Voyage %s: rows %d-%d highlighted with color #%s",
            voyage.get("voyage_no", v_idx + 1),
            excel_start, excel_end, color_hex,
        )

    # -- 2. Apply RED BOLD font to all used columns (all data rows) --
    for row_num in range(2, total_rows + 1):  # skip header row
        for col_num in used_excel_cols:
            if col_num <= total_cols:
                cell = ws.cell(row=row_num, column=col_num)
                # Preserve existing font properties but override color and bold
                old_font = cell.font
                cell.font = Font(
                    name=old_font.name,
                    size=old_font.size,
                    italic=old_font.italic,
                    underline=old_font.underline,
                    strike=old_font.strike,
                    color="FF0000",
                    bold=True,
                )

    # -- 3. Add a legend sheet --
    if "Voyage Legend" in wb.sheetnames:
        del wb["Voyage Legend"]
    legend = wb.create_sheet("Voyage Legend")
    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 40
    legend.column_dimensions["C"].width = 20
    legend.column_dimensions["D"].width = 20

    legend.cell(row=1, column=1, value="Voyage").font = Font(bold=True)
    legend.cell(row=1, column=2, value="Type").font = Font(bold=True)
    legend.cell(row=1, column=3, value="Departure").font = Font(bold=True)
    legend.cell(row=1, column=4, value="Arrival").font = Font(bold=True)
    legend.cell(row=1, column=5, value="Color").font = Font(bold=True)

    for v_idx, voyage in enumerate(voyages):
        row = v_idx + 2
        color_hex = VOYAGE_COLORS[v_idx % len(VOYAGE_COLORS)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        legend.cell(row=row, column=1, value=voyage.get("voyage_no", f"Voyage {v_idx + 1}"))
        legend.cell(row=row, column=2, value=voyage.get("voyage_type", ""))
        legend.cell(row=row, column=3, value=str(voyage.get("dep_datetime", ""))[:16])
        legend.cell(row=row, column=4, value=str(voyage.get("arr_datetime", ""))[:16])
        legend.cell(row=row, column=5, value=f"#{color_hex}")

        for col in range(1, 6):
            legend.cell(row=row, column=col).fill = fill

    # Note about red bold columns
    note_row = len(voyages) + 3
    legend.cell(row=note_row, column=1, value="Note:").font = Font(bold=True)
    legend.cell(row=note_row, column=2,
                value="Columns with RED BOLD font are used for report generation.")
    legend.cell(row=note_row, column=2).font = Font(color="FF0000", bold=True)

    # -- Save --
    wb.save(str(output_path))
    logger.info("Highlighted report saved to: %s", output_path)
